#!/usr/bin/env python3
"""
Inspect the structure of the Excel file to understand sheets and columns
"""

import pandas as pd
import openpyxl

EXCEL_FILE = "/var/www/html/circuitinfo/Discount Tire Master Circuit List.xlsx"

print("Inspecting Excel file structure...")

# First, check all sheet names
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
print(f"\nSheet names: {wb.sheetnames}")

# Check each sheet
for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: '{sheet_name}' ---")
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, nrows=5)
        print(f"Shape: {df.shape}")
        print(f"Columns ({len(df.columns)}): {list(df.columns)[:10]}...")  # First 10 columns
        
        # If this sheet has enough columns, check columns A, Q-X
        if len(df.columns) >= 24:
            print(f"\nColumn A (index 0): {df.columns[0]}")
            print(f"Column Q (index 16): {df.columns[16]}")
            print(f"Column R (index 17): {df.columns[17]}")
            print(f"Column S (index 18): {df.columns[18]}")
            print(f"Column T (index 19): {df.columns[19]}")
            print(f"Column U (index 20): {df.columns[20]}")
            print(f"Column V (index 21): {df.columns[21]}")
            print(f"Column W (index 22): {df.columns[22]}")
            print(f"Column X (index 23): {df.columns[23]}")
    except Exception as e:
        print(f"Error reading sheet: {e}")

wb.close()